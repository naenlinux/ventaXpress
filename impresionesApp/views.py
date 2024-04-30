from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML, CSS
from ventaApp.models import Ventas, PedidosDetalle, Pedidos
from empresaApp.models import Empresa, Sucursal
from inventarioApp.models import Almacen, Compras
from mantenedoresApp.models import Productos, UnidadMedida
from pprint import pprint
from openpyxl import Workbook
import io, os, qrcode, base64
from openpyxl.styles import colors, Font, Color, PatternFill, Border, Side
from datetime import date, datetime
from mantenedoresApp.models import Productos, Proveedores, Categorias
from .serializers import ProveedoresSerializer, CategoriasSerializer, ProductosSerializer, PedidosSerializer, ComprasSerializer, SucursalSerializer
from rest_framework import viewsets
from rest_framework.response import Response
from django.db.models import Sum, F
from django.db.models.functions import ExtractMonth, TruncMonth
from django.utils.timezone import now
from ventaApp.serializers import VentasSerializer
from io import BytesIO


def imprimirComprobante(request, idVenta, idprint):
    venta = Ventas.objects.select_related('idPedido','tipComprobante').get(pk=idVenta)
    empresa = Empresa.objects.first()
    detalles = PedidosDetalle.objects.filter(idPedido=venta.idPedido)
    importe_total = venta.idPedido.subtotal + venta.idPedido.igv_total

    imporSinDscto = sum(detalle.precio * detalle.cantidad for detalle in detalles)
    

    detalles_array = []

    QR = generarQR(f"{empresa.ruc}|{venta.tipComprobante.nombre}|{venta.serie}|{venta.numComprobante}|{venta.idPedido.igv_total}|{importe_total}|{venta.fecha}")

    for detalle in detalles:
        costo_bruto = round(detalle.precio * detalle.cantidad,2)
        detalle.costo_bruto = "{:,.2f}".format(costo_bruto)
    dscto_total = venta.idPedido.dscto_total
    context = {
        'venta': venta,
        'empresa': empresa,
        'detalles': detalles,
        'importe_total': "{:,.2f}".format(importe_total),
        'imporSinDscto':"{:,.2f}".format(imporSinDscto),
        'dscto_total':"{:,.2f}".format(dscto_total),
        'totaletra':numero_a_moneda_sunat(importe_total),
        'qr':QR
    }
    
    if(idprint == 1): # IMPRESORA A4
        template = get_template('comproVentaA4.html')
        
        html_string = template.render(context)
        
        # Convierte el HTML en un documento PDF
        html = HTML(string=html_string)
        html.page_size = 'A4'
        
        pdf_file = html.write_pdf()

        # Crea una respuesta HTTP con el PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="comproVentaA4.pdf"'
    else: #TICKETERA 80mm
        template = get_template('comproVentaTick.html')
        html_string = template.render(context)
        
        # Convierte el HTML en un documento PDF
        html = HTML(string=html_string)
        
        # Configura el tamaño de la página para la ticketera (80mm de ancho)
        altosize = 170 + len(detalles) * 5
        alto = f"{altosize}mm"
        html.write_pdf(target='/tmp/comproVentaTick.pdf', presentational_hints=True,  stylesheets=[CSS(string='@page { size: 80mm '+alto+'; margin: 0; }')])
        
        # Abre el archivo PDF y lee su contenido
        with open('/tmp/comproVentaTick.pdf', 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="comproVentaTick.pdf"'
        
        # Elimina el archivo PDF temporal
        os.remove('/tmp/comproVentaTick.pdf')

    return response

def export_to_excel(request, idcategoria):
    if(idcategoria > 0):
        almacen = Almacen.objects.select_related('idProducto__unidadMedida','idProducto__categoria').filter(activo=1,idProducto__categoria=idcategoria)
    else:
        almacen = Almacen.objects.select_related('idProducto__unidadMedida','idProducto__categoria').filter(activo=1)

    wb = Workbook()
    ws = wb.active
    fecha_hoy = date.today()
    fecha_datetime = datetime.strptime(str(fecha_hoy),"%Y-%m-%d")
    fecha_forma = fecha_datetime.strftime("%d/%m/%y")

   
    hora = datetime.now().time()
    objeto_tiempo = datetime.strptime(str(hora), "%H:%M:%S.%f")
    hora_form = objeto_tiempo.strftime("%H:%M")
    ws['A2'] = 'REPORTE DE ALMACEN ' + fecha_forma +' '+str(hora_form)
    ws['A2'].fill = PatternFill(start_color='00ffaf00', end_color='00ffaf00',fill_type='solid')
    ws['A2'].font = Font(bold=True)

    ws.merge_cells('A2:E2')
    ws['A4'] = 'N°'
    ws['B4'] = 'PRODUCTO'
    ws['C4'] = 'CATEGORIA'
    ws['D4'] = 'STOCK X U/M'
    ws['E4'] = 'STOCK X UNIDAD'

    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws['E4'].font = Font(bold=True)

    ws['A4'].fill = PatternFill(start_color='0000baff', end_color='0000baff',fill_type='solid')
    ws['B4'].fill = PatternFill(start_color='0000baff', end_color='0000baff',fill_type='solid')
    ws['C4'].fill = PatternFill(start_color='0000baff', end_color='0000baff',fill_type='solid')
    ws['D4'].fill = PatternFill(start_color='0000baff', end_color='0000baff',fill_type='solid')
    ws['E4'].fill = PatternFill(start_color='0000baff', end_color='0000baff',fill_type='solid')

    cont = 5
    numerador = 1

    for almace in almacen:
        ws.cell(row = cont, column = 1).value = numerador
        ws.cell(row = cont, column = 2).value = almace.idProducto.nombre
        ws.cell(row = cont, column = 3).value = almace.idProducto.categoria.nombre
        ws.cell(row = cont, column = 4).value = str(almace.cantidad) +' '+almace.idProducto.unidadMedida.nombre
        ws.cell(row = cont, column = 5).value = almace.total
        cont += 1
        numerador += 1
    for row in ws.iter_rows(min_row=4, max_row=cont-1, min_col=1, max_col=5):
        for cell in row:
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border
    
    ancho_columnas = ['B','C','D','E']
    ancho_personalizado = 20

    for columna in ancho_columnas:
        ws.column_dimensions[columna].width = ancho_personalizado

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ReporteAlmacenExcel.xlsx"'

    # Guarda el libro de trabajo en un objeto BytesIO en lugar de disco
    #excel_file = io.BytesIO()
    wb.save(response)
    #response.write(excel_file.getvalue())

    return response

def excelReporteVentas(request,fini, ffin):
    queryset = Ventas.objects.filter(fecha__range=(fini,ffin)).exclude(activo=0).order_by('id')
    serializer_class = VentasSerializer(queryset,many=True)
    
    data = serializer_class.data

    wb = Workbook()
    ws = wb.active
    fecha_hoy = date.today()
    fecha_datetime = datetime.strptime(str(fecha_hoy),"%Y-%m-%d")
    fecha_forma = fecha_datetime.strftime("%d/%m/%y")

   
    hora = datetime.now().time()
    objeto_tiempo = datetime.strptime(str(hora), "%H:%M:%S.%f")
    hora_form = objeto_tiempo.strftime("%H:%M")
    fini_date = datetime.strptime(fini, '%Y-%m-%d')
    ffin_date = datetime.strptime(ffin, '%Y-%m-%d')
    ws['B2'] = 'REPORTE DE VENTAS  DEL ' + fini_date.strftime('%d/%m/%Y') +' AL '+ffin_date.strftime('%d/%m/%Y')
    ws['B2'].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00',fill_type='solid')
    ws['B2'].font = Font(bold=True)

    ws.merge_cells('B2:D2')

    ws['A4'] = 'ESTADO'
    ws['B4'] = 'COMPROBANTE'
    ws['C4'] = 'SERIE'
    ws['D4'] = 'FECHA DE VENTA'
    ws['E4'] = 'SUBTOTAL'
    ws['F4'] = 'METODO PAGO'
    ws['G4'] = 'CLIENTE'
    ws['H4'] = 'SUCURSAL'
    ws['I4'] = 'VENDEDOR'
    ws['J4'] = 'COBRADO POR'
    ws['K4'] = 'NUMERO DE PEDIDO'
    ws['L4'] = 'FECHA DE PEDIDO'


    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)
    ws['E4'].font = Font(bold=True)
    ws['F4'].font = Font(bold=True)
    ws['G4'].font = Font(bold=True)
    ws['H4'].font = Font(bold=True)
    ws['I4'].font = Font(bold=True)
    ws['J4'].font = Font(bold=True)
    ws['K4'].font = Font(bold=True)
    ws['L4'].font = Font(bold=True)

    ws['A4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['B4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['C4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['D4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['E4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['F4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['G4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['H4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['I4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['J4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['K4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['L4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')

    cont = 5

    for venta in data:
        if(venta['activo'] == 1):
            ws.cell(row = cont, column = 1).value = 'Activo'
            signo = ''
        else:
            ws.cell(row = cont, column = 1).value = 'Anulado'
            signo = '-'
        ws.cell(row = cont, column = 2).value = venta['comprobante']
        ws.cell(row = cont, column = 3).value = venta['serie'] + venta['numComprobante']
        ws.cell(row = cont, column = 4).value = formatear_fecha(venta['fecha'])
        ws.cell(row = cont, column = 5).value = signo+venta['subtotal']
        ws.cell(row = cont, column = 6).value = venta['metodoPago']
        ws.cell(row = cont, column = 7).value = venta['cliente']
        ws.cell(row = cont, column = 8).value = venta['sucursal']
        ws.cell(row = cont, column = 9).value = venta['vendedor']
        ws.cell(row = cont, column = 10).value = venta['cobrador']
        ws.cell(row = cont, column = 11).value = venta['numPedido']
        ws.cell(row = cont, column = 12).value = formatear_fecha(venta['fechaPedido'])
        cont += 1

    for row in ws.iter_rows(min_row=4, max_row=cont-1, min_col=1, max_col=12):
        for cell in row:
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border
    
    ancho_columnas = ['A','B','C','D','E','F','G','H','I','J','K','L']
    ancho_personalizado = 20

    for columna in ancho_columnas:
        ws.column_dimensions[columna].width = ancho_personalizado

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ReporteVentas.xlsx"'

    # Guarda el libro de trabajo en un objeto BytesIO en lugar de disco
    #excel_file = io.BytesIO()
    wb.save(response)
    #response.write(excel_file.getvalue())

    return response

class countProveedoresviewSet(viewsets.ViewSet):
    serializer_class = ProveedoresSerializer

    def get_queryset(self):
        return Proveedores.objects.filter(activo=True)

    def list(self, request):
        queryset = self.get_queryset()
        conteo = queryset.count()
        return Response({'countProveedores': conteo})

class countProductoviewSet(viewsets.ViewSet):
    serializer_class = ProductosSerializer

    def get_queryset(self):
        return Productos.objects.filter(activo=True)

    def list(self, request):
        queryset = self.get_queryset()
        conteo = queryset.count()
        return Response({'countProductos': conteo})

class countCategoriaviewSet(viewsets.ViewSet):
    serializer_class = CategoriasSerializer

    def get_queryset(self):
        return Categorias.objects.filter(activo=True)

    def list(self, request):
        queryset = self.get_queryset()
        conteo = queryset.count()
        return Response({'countCategorias': conteo})

class sumaVentasViewset(viewsets.ViewSet):
    serializer_class = PedidosSerializer   

    def list(self, request):
        suma_ventas = Pedidos.objects.filter(activo=True, estado='Pagado').aggregate(suma_ventas=Sum('subtotal') + Sum('igv_total') + Sum('dscto_total'))
        return Response({'sumaVentas': suma_ventas['suma_ventas']})

class sumaComprasViewset(viewsets.ViewSet):
    serializer_class = ComprasSerializer   

    def list(self, request):
        suma_compras = Compras.objects.filter(activo=True).aggregate(suma_compras=Sum('compraTotal'))
        return Response({'sumaCompras': suma_compras['suma_compras']})

class countSucursalviewSet(viewsets.ViewSet):
    serializer_class = SucursalSerializer

    def get_queryset(self):
        return Sucursal.objects.filter(activo=True)

    def list(self, request):
        queryset = self.get_queryset()
        conteo = queryset.count()
        return Response({'countSucursal': conteo})

class ventaxMesViewsets(viewsets.ViewSet):
    serializer_class = PedidosSerializer

    def list(self, request):
        anio_actual = now().year
        ventas_por_mes = Pedidos.objects.filter(
            activo=True, estado = 'Pagado', fecha__year = anio_actual
        ).annotate(
        mes=ExtractMonth('fecha')
        ).values('mes').annotate(
            venta_total=Sum(F('subtotal') + F('igv_total') + F('dscto_total') )
        ).order_by('mes')

        return Response(ventas_por_mes)



MONEDA_SINGULAR = 'NUEVO SOL'
MONEDA_PLURAL = 'NUEVOS SOLES'

CENTIMOS_SINGULAR = 'CENTIMO'
CENTIMOS_PLURAL = 'CENTIMOS'

MAX_NUMERO = 999999999999

UNIDADES = (
    'CERO',
    'UNO',
    'DOS',
    'TRES',
    'CUATRO',
    'CINCO',
    'SEIS',
    'SIETE',
    'OCHO',
    'NUEVE'
)

DECENAS = (
    'DIEZ',
    'ONCE',
    'DOCE',
    'TRECE',
    'CATORCE',
    'QUINCE',
    'DIECISEIS',
    'DIECISIETE',
    'DIECIOCHO',
    'DIECINUEVE'
)

DIEZ_DIEZ = (
    'CERO',
    'DIEZ',
    'VEINTE',
    'TREINTA',
    'CUARENTA',
    'CINCUENTA',
    'SESENTA',
    'SETENTA',
    'OCHENTA',
    'NOVENTA'
)

CIENTOS = (
    '_',
    'CIENTO',
    'DOSCIENTOS',
    'TRESCIENTOS',
    'CUATROSCIENTOS',
    'QUINIENTOS',
    'SEISCIENTOS',
    'SETECIENTOS',
    'OCHOCIENTOS',
    'NOVECIENTOS'
)


def numero_a_letras(numero):
    numero_entero = int(numero)
    if numero_entero > MAX_NUMERO:
        raise OverflowError('Número demasiado alto')
    if numero_entero < 0:
        negativo_letras = numero_a_letras(abs(numero))
        return f"MENOS {negativo_letras}"
    letras_decimal = ''
    parte_decimal = int(round((abs(numero) - abs(numero_entero)) * 100))
    if parte_decimal > 9:
        letras_decimal = f"PUNTO {numero_a_letras(parte_decimal)}"
    elif parte_decimal > 0:
        letras_decimal = f"PUNTO CERO {numero_a_letras(parte_decimal)}"
    if numero_entero <= 99:
        resultado = leer_decenas(numero_entero)
    elif numero_entero <= 999:
        resultado = leer_centenas(numero_entero)
    elif numero_entero <= 999999:
        resultado = leer_miles(numero_entero)
    elif numero_entero <= 999999999:
        resultado = leer_millones(numero_entero)
    else:
        resultado = leer_millardos(numero_entero)
    resultado = resultado.replace('UNO MIL', 'UN MIL')
    resultado = resultado.strip()
    resultado = resultado.replace(' _ ', ' ')
    resultado = resultado.replace('  ', ' ')
    if parte_decimal > 0:
        resultado = f"{resultado} {letras_decimal}"
    return resultado


def numero_a_moneda(numero):
    numero_entero = int(numero)
    parte_decimal = int(round((abs(numero) - abs(numero_entero)) * 100))
    centimos = CENTIMOS_SINGULAR if parte_decimal == 1 else CENTIMOS_PLURAL
    moneda = MONEDA_SINGULAR if numero_entero == 1 else MONEDA_PLURAL
    letras = numero_a_letras(numero_entero)
    letras = letras.replace('UNO', 'UN')
    aux_decimal = numero_a_letras(parte_decimal).replace('UNO', 'UN')
    letras_decimal = f"con {aux_decimal} {centimos}"
    letras = f"{letras} {letras_decimal} {moneda}"
    return letras


def leer_decenas(numero):
    if numero < 10:
        return UNIDADES[numero]
    decena, unidad = divmod(numero, 10)
    if numero <= 19:
        resultado = DECENAS[unidad]
    elif numero <= 29:
        resultado = f"VEINTI{UNIDADES[unidad]}"
    else:
        resultado = DIEZ_DIEZ[decena]
        if unidad > 0:
            resultado = f"{resultado} y {UNIDADES[unidad]}"
    return resultado


def leer_centenas(numero):
    centena, decena = divmod(numero, 100)
    if numero == 0:
        resultado = 'CIEN'
    else:
        resultado = CIENTOS[centena]
        if decena > 0:
            decena_letras = leer_decenas(decena)
            resultado = f"{resultado} {decena_letras}"
    return resultado


def leer_miles(numero):
    millar, centena = divmod(numero, 1000)
    resultado = ''
    if millar == 1:
        resultado = ''
    if (millar >= 2) and (millar <= 9):
        resultado = UNIDADES[millar]
    elif (millar >= 10) and (millar <= 99):
        resultado = leer_decenas(millar)
    elif (millar >= 100) and (millar <= 999):
        resultado = leer_centenas(millar)
    resultado = f"{resultado} MIL"
    if centena > 0:
        centena_letras = leer_centenas(centena)
        resultado = f"{resultado} {centena_letras}"
    return resultado.strip()


def leer_millones(numero):
    millon, millar = divmod(numero, 1000000)
    resultado = ''
    if millon == 1:
        resultado = ' UN MILLON '
    if (millon >= 2) and (millon <= 9):
        resultado = UNIDADES[millon]
    elif (millon >= 10) and (millon <= 99):
        resultado = leer_decenas(millon)
    elif (millon >= 100) and (millon <= 999):
        resultado = leer_centenas(millon)
    if millon > 1:
        resultado = f"{resultado} MILLONES"
    if (millar > 0) and (millar <= 999):
        centena_letras = leer_centenas(millar)
        resultado = f"{resultado} {centena_letras}"
    elif (millar >= 1000) and (millar <= 999999):
        miles_letras = leer_miles(millar)
        resultado = f"{resultado} {miles_letras}"
    return resultado


def leer_millardos(numero):
    millardo, millon = divmod(numero, 1000000)
    miles_letras = leer_miles(millardo)
    millones_letras = leer_millones(millon)
    return f"{miles_letras} MILLONES {millones_letras}"


def numero_a_moneda_sunat(numero):
    numero_entero = int(numero)
    parte_decimal = int(round((abs(numero) - abs(numero_entero)) * 100))
    moneda = MONEDA_SINGULAR if numero_entero == 1 else MONEDA_PLURAL

    letras = numero_a_letras(numero_entero)
    letras = letras.replace('UNO', 'UN')
    letras = f"{letras} Y {parte_decimal}/100 {moneda}"
    return letras

def generarQR(datos):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(datos)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")

    # Convierte la imagen del QR a base64
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()
    buffer.close()

    return qr_base64